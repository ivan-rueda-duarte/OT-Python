from openpyxl import load_workbook
import datetime
from datetime import date


class Tarea:
    
    def __init__(self, nombre, fecha_inicio, fecha_fin, grupo_de_tarea, actividad, personas, cargo, duracion):
        self.nombre = nombre
        self.fecha_inicio = fecha_inicio
        self.fecha_fin = fecha_fin
        self.grupo_de_tarea = grupo_de_tarea
        self.actividad = actividad
        self.personas = personas
        self.cargo = cargo
        self.duracion = duracion

weekDays = ("Lunes","Martes","Miercoles","Jueves","Viernes","Sabado","Domingo")

#Para cargar el archivo
gantt = load_workbook('gantt.xlsx')

#para crear el objecto hoja y manipularlos
hoja = gantt.active

 
row_test= []    #se crea una lista para utilizarla luego
columna_test = hoja['A']    #selecciona la primera columna para saber la longitud
lista_tarea = []    #se crea una lista para utilizarla luego
i=2 #Desde que fila comienzan las tareas en el Gantt


#for para la cantidad de tareas
for y in columna_test:
    row_test = hoja[i]
    

    #Se distribuye la info de la tarea en los correspondientes valores
    for x in row_test:
        nombre = row_test[0]    #colocar nombre
        fecha_inicio = row_test[1]  #colocar fecha de inicio
        fecha_fin= row_test[2]      #colocar fecha de fin
        grupo_de_tarea = row_test[3]     #buscar manera más elegante de colocarlo
        actividad = row_test[4]     #Coloca la actividad a realizar
        personas = row_test[5]      #Coloca las personas a realizar la actividad
        cargo = row_test[6]         #El cargo de las personas 
        duracion = row_test[7]      #la duracion de la actividas


    if nombre.value == None:
        break
    #se crea una objecto que almacene los datos de la tarea
    tarea_index = Tarea(nombre.value, fecha_inicio.value, fecha_fin.value, grupo_de_tarea.value, actividad.value, personas.value, cargo.value, duracion.value) 
    #Se agrega las tareas a una lista para utilizarla luego
    lista_tarea.append(tarea_index)
    i += 1  #incrementa en uno para buscar la siguiente fila y disernir los datos
    
    


#hacer un for loop para cambiar los documentos que se abren
t=1
for x in lista_tarea:

    
    wb2 = load_workbook('OT-0-Plantilla.xlsx')    #para cargar el documento a trabajar, 
                            #la t es para saber el numero a documento a cargar

    hoja_out = wb2.active   #activar la hoja
    
    
    hoja_out["E1"] = "Orden de trabajo Nº 0000" + str(t)        #orden de trabajo numero..
    
    hoja_out["E5"] = lista_tarea[t-1].grupo_de_tarea       #Nombre del grupo de tarea

    hoja_out["E6"]=lista_tarea[t-1].nombre #Nombre de la tarea

    hoja_out["I5"] = weekDays[lista_tarea[t-1].fecha_inicio.weekday()]        #Día de la semana

    hoja_out["I6"] =  lista_tarea[t-1].fecha_inicio      #fecha de inicio

    hoja_out["C12"] =  lista_tarea[t-1].actividad       #actividad a desarrollar

    hoja_out["G12"] =  lista_tarea[t-1].personas        #personas 

    hoja_out["H12"] =  lista_tarea[t-1].cargo           #cargo de las personas

    hoja_out["J12"] =  lista_tarea[t-1].duracion        #duracion de la actividad en minutos



    wb2.save('OT-'+ str(t) +'.xlsx')   #guardar el documento

    t += 1



